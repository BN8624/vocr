# 원본표 Excel 시트 생성 계약을 검증하는 테스트
from __future__ import annotations

import json
import sys
import tempfile
from datetime import date, datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import load_workbook

from src.excel_exporter import _infer_years_for_rows, _parse_statement_period_text, export_excel
from src.validator import ValidationOutput


def main() -> int:
    with tempfile.TemporaryDirectory() as temp_dir:
        root = Path(temp_dir)
        merged_dir = root / "merged"
        merged_dir.mkdir()
        validated_path = merged_dir / "transactions_validated.jsonl"
        source_rows_path = merged_dir / "rows_merged.jsonl"
        issues_path = merged_dir / "validation_issues.json"
        summary_path = merged_dir / "validation_summary.json"

        _write_jsonl(validated_path, [_normalized_row()])
        _write_jsonl(
            source_rows_path,
            [
                _source_row(
                    page=1,
                    local_row_index=1,
                    header=["이용일", "이용카드", "가맹점명", "결제원금"],
                    cells=["01.01", "본인카드", "상점A", "1,000", "남는값"],
                ),
                _source_row(
                    page=1,
                    local_row_index=2,
                    header=["이용일", "이용카드", "가맹점명", "결제원금"],
                    cells=["01.02", "가족카드"],
                ),
                _source_row(
                    page=1,
                    local_row_index=3,
                    header=["구분", "금액텍스트", "금액"],
                    cells=["총 합계 결제원금", "3,000", "3000"],
                    row_type="total",
                ),
            ],
        )
        summary = {"checksum": {"status": "no_source_total", "source_total_candidates": []}}
        summary_path.write_text(json.dumps(summary, ensure_ascii=False), encoding="utf-8")
        issues_path.write_text("{}", encoding="utf-8")

        output = export_excel(
            validation_output=ValidationOutput(
                validated_transactions_path=validated_path,
                issues_path=issues_path,
                summary_path=summary_path,
                transaction_count=1,
                row_issue_count=0,
                issue_row_count=0,
                checksum_status="no_source_total",
                checksum_difference=None,
                summary=summary,
            ),
            output_dir=root,
            source_rows_path=source_rows_path,
        )

        assert output is not None
        workbook = load_workbook(output.workbook_path, read_only=True)
        try:
            assert workbook.sheetnames == ["원본표", "원본표_개발자", "전체명세_정규화", "검산", "원본셀", "추가필드", "확인필요"]
            sheet = workbook["원본표"]
            headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
            assert headers[:4] == ["이용일", "이용카드", "가맹점명", "결제원금"]
            assert "page" not in headers
            assert "extra_col_1" not in headers

            first = [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))]
            second = [cell.value for cell in next(sheet.iter_rows(min_row=3, max_row=3))]
            assert _date_text(first[0]) == f"{date.today().year}-01-01"
            assert first[1:4] == ["본인카드", "상점A", 1000]
            assert _date_text(second[0]) == f"{date.today().year}-01-02"
            assert second[1:4] == ["가족카드", None, None]
            assert sheet.max_row == 3

            dev_sheet = workbook["원본표_개발자"]
            dev_headers = [cell.value for cell in next(dev_sheet.iter_rows(min_row=1, max_row=1))]
            assert dev_headers[:8] == [
                "page",
                "chunk_id",
                "local_row_index",
                "row_type",
                "이용일",
                "이용카드",
                "가맹점명",
                "결제원금",
            ]
            assert "extra_col_1" in dev_headers
            dev_first = [cell.value for cell in next(dev_sheet.iter_rows(min_row=2, max_row=2))]
            assert dev_first[4:8] == ["01.01", "본인카드", "상점A", "1,000"]
            assert dev_first[dev_headers.index("extra_col_1")] == "남는값"
            dev_third = [cell.value for cell in next(dev_sheet.iter_rows(min_row=4, max_row=4))]
            assert dev_third[3] == "total"
            assert dev_third[dev_headers.index("구분")] == "총 합계 결제원금"
            assert dev_third[dev_headers.index("금액텍스트")] == "3,000"
            assert dev_third[dev_headers.index("금액")] == "3000"
            assert dev_sheet.max_row == 4
        finally:
            workbook.close()

    period = _parse_statement_period_text("[일시불 할부] 2024.12.,09 2025.01.08")
    assert period is not None
    year_rows = [
        _source_row(1, 1, ["이용일"], ["1211"]),
        _source_row(1, 2, ["이용일"], ["0102"]),
    ]
    years = _infer_years_for_rows(year_rows, [0], period)
    assert years[(0, 0)] == 2024
    assert years[(1, 0)] == 2025

    print("original table export test passed")
    return 0


def _source_row(
    page: int,
    local_row_index: int,
    header: list[str],
    cells: list[str],
    row_type: str | None = None,
) -> dict[str, object]:
    return {
        **({"row_type": row_type} if row_type else {}),
        "source": {
            "file": "sample.pdf",
            "page": page,
            "chunk_id": f"page_{page:03d}_chunk_01",
            "local_row_index": local_row_index,
        },
        "raw": {"header": header, "cells": cells},
    }


def _normalized_row() -> dict[str, object]:
    return {
        "source": {"file": "sample.pdf", "page": 1, "chunk_id": "page_001_chunk_01", "local_row_index": 1},
        "raw": {"header": ["이용일", "가맹점명"], "cells": ["01.01", "상점A"]},
        "transaction": {"date": "01.01", "merchant": "상점A", "amount": 1000, "billing_amount": 1000},
        "quality": {"needs_review": False, "review_reason": ""},
        "validation": {"needs_review": False, "issues": []},
        "extra_fields": {},
    }


def _date_text(value: object) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def _write_jsonl(path: Path, rows: list[dict[str, object]]) -> None:
    with path.open("w", encoding="utf-8") as file:
        for row in rows:
            file.write(json.dumps(row, ensure_ascii=False) + "\n")


if __name__ == "__main__":
    raise SystemExit(main())
