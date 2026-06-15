from __future__ import annotations

import argparse
import json
import re
import subprocess
import sys
from datetime import datetime
from pathlib import Path
from typing import Any


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Run sample PDFs through the local pipeline and write a regression report."
    )
    parser.add_argument("--samples-dir", default="견본", help="Folder containing *_N.pdf files.")
    parser.add_argument("--output", default="output/regression_samples", help="Regression output folder.")
    parser.add_argument("--with-vision", action="store_true", help="Call Vision API instead of dry-run mode.")
    parser.add_argument("--no-force", action="store_true", help="Reuse rendered page/chunk images when present.")
    parser.add_argument("--limit", type=int, default=0, help="Limit number of sample PDFs. 0 means all.")
    parser.add_argument("--pages", type=int, default=0, help="Run only samples whose filename page count matches this value. 0 means all.")
    parser.add_argument("--issuer", default="", help="Run only samples whose filename starts with this issuer/card-company name.")
    args = parser.parse_args()

    root = Path(__file__).resolve().parents[1]
    samples_dir = (root / args.samples_dir).resolve()
    output_root = (root / args.output).resolve()
    output_root.mkdir(parents=True, exist_ok=True)

    samples = find_sample_pdfs(samples_dir)
    if args.pages > 0:
        samples = [path for path in samples if expected_page_count(path) == args.pages]
    if args.issuer:
        samples = [path for path in samples if sample_issuer(path) == args.issuer]
    if args.limit > 0:
        samples = samples[: args.limit]
    if not samples:
        raise FileNotFoundError(f"No sample PDFs found: {samples_dir}")

    results = [
        run_sample(
            root=root,
            pdf_path=pdf_path,
            output_root=output_root,
            dry_run=not args.with_vision,
            force=not args.no_force,
        )
        for pdf_path in samples
    ]
    report = {
        "schema_version": "1.0",
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "mode": "vision" if args.with_vision else "dry_run",
        "sample_count": len(results),
        "pass_count": sum(1 for result in results if result["status"] == "PASS"),
        "fail_count": sum(1 for result in results if result["status"] != "PASS"),
        "results": results,
    }

    json_path = output_root / "sample_regression_report.json"
    md_path = output_root / "sample_regression_report.md"
    json_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    md_path.write_text(markdown_report(report, root), encoding="utf-8")

    print(f"Report JSON: {json_path}")
    print(f"Report Markdown: {md_path}")
    for result in results:
        print(
            "\t".join(
                [
                    str(result["sample"]),
                    f"pages={result['page_count']}/{result['expected_pages']}",
                    f"chunks={result['chunk_count']}",
                    f"vision={result['vision_ok']}/{result['vision_errors']}",
                    str(result["checksum_status"]),
                    str(result["status"]),
                ]
            )
        )

    return 0 if report["fail_count"] == 0 else 1


def find_sample_pdfs(samples_dir: Path) -> list[Path]:
    if not samples_dir.exists():
        raise FileNotFoundError(f"Sample folder not found: {samples_dir}")
    return sorted(samples_dir.glob("*.pdf"), key=lambda path: path.name)


def expected_page_count(pdf_path: Path) -> int:
    match = re.search(r"_(\d+)$", pdf_path.stem)
    if not match:
        raise ValueError(f"Sample filename must end with _<page_count>: {pdf_path}")
    return int(match.group(1))


def sample_issuer(pdf_path: Path) -> str:
    match = re.match(r"^(.+)_(\d+)$", pdf_path.stem)
    return match.group(1) if match else pdf_path.stem


def run_sample(root: Path, pdf_path: Path, output_root: Path, dry_run: bool, force: bool) -> dict[str, Any]:
    output_dir = output_root / pdf_path.stem
    command = [
        sys.executable,
        str(root / "main.py"),
        "--input",
        str(pdf_path),
        "--output",
        str(output_dir),
    ]
    if dry_run:
        command.append("--dry-run")
    if force:
        command.append("--force")

    completed = subprocess.run(
        command,
        cwd=root,
        text=True,
        capture_output=True,
        encoding="utf-8",
        errors="replace",
    )
    summary_path = output_dir / "summary.json"
    summary = read_json_object(summary_path)
    body_chunk_count = manifest_count(output_dir / "chunks" / "chunks_manifest.json")
    total_chunk_count = manifest_count(output_dir / "total_chunks" / "total_chunks_manifest.json")
    validation_summary = read_json_object(output_dir / "merged" / "validation_summary.json")
    checksum = validation_summary.get("checksum", {}) if isinstance(validation_summary.get("checksum"), dict) else {}
    normalization = read_json_object(output_dir / "merged" / "normalization_summary.json")
    merge_summary = read_json_object(output_dir / "merged" / "merge_summary.json")
    source_rows_path = output_dir / "merged" / "rows_merged.jsonl"
    original_source = inspect_original_source(source_rows_path)

    expected_pages = expected_page_count(pdf_path)
    page_count = int(summary.get("page_count", 0) or 0)
    chunk_count = int(summary.get("chunk_count", 0) or 0)
    review_html = Path(str(summary.get("review_html", output_dir / "review.html")))
    excel_path = Path(str(summary.get("excel_path", output_dir / "result.xlsx")))
    excel_check = inspect_excel(excel_path, original_source)
    checks = [
        completed.returncode == 0,
        page_count == expected_pages,
        chunk_count > 0,
        body_chunk_count > 0,
        total_chunk_count == expected_pages,
        review_html.exists(),
        excel_check["exists"],
        excel_check["original_table_first"],
        excel_check["original_table_non_empty"],
        excel_check["original_row_coverage_ok"],
        excel_check["original_cell_coverage_ok"],
        excel_check["normalized_sheet_exists"],
        excel_check["checksum_sheet_exists"],
    ]
    status = "PASS" if all(checks) else "FAIL"

    return {
        "sample": str(pdf_path.relative_to(root)),
        "expected_pages": expected_pages,
        "page_count": page_count,
        "chunk_count": chunk_count,
        "body_chunk_count": body_chunk_count,
        "total_chunk_count": total_chunk_count,
        "phase": summary.get("phase", ""),
        "vision_ok": int(summary.get("vision_ok", 0) or 0),
        "vision_errors": int(summary.get("vision_errors", 0) or 0),
        "raw_row_count": int(summary.get("raw_row_count", 0) or merge_summary.get("raw_row_count", 0) or 0),
        "duplicate_group_count": int(summary.get("duplicate_group_count", 0) or merge_summary.get("duplicate_group_count", 0) or 0),
        "transaction_count": int(summary.get("transaction_count", 0) or normalization.get("transaction_count", 0) or 0),
        "validation_issue_row_count": int(summary.get("validation_issue_row_count", 0) or 0),
        "checksum_status": str(summary.get("checksum_status") or checksum.get("status") or "not_run"),
        "source_total_candidate_count": len(checksum.get("source_total_candidates", [])) if isinstance(checksum, dict) else 0,
        "review_html": str(review_html),
        "excel_exists": excel_path.exists(),
        "excel_sheets": excel_check["sheet_names"],
        "original_table_first": excel_check["original_table_first"],
        "original_table_non_empty": excel_check["original_table_non_empty"],
        "original_source_row_count": original_source["row_count"],
        "original_source_nonblank_cell_count": original_source["nonblank_cell_count"],
        "original_table_row_count": excel_check["original_table_row_count"],
        "original_table_column_count": excel_check["original_table_column_count"],
        "original_table_nonblank_cell_count": excel_check["original_table_nonblank_cell_count"],
        "original_row_coverage_rate": excel_check["original_row_coverage_rate"],
        "original_cell_coverage_rate": excel_check["original_cell_coverage_rate"],
        "original_row_coverage_ok": excel_check["original_row_coverage_ok"],
        "original_cell_coverage_ok": excel_check["original_cell_coverage_ok"],
        "normalized_sheet_exists": excel_check["normalized_sheet_exists"],
        "checksum_sheet_exists": excel_check["checksum_sheet_exists"],
        "returncode": completed.returncode,
        "status": status,
        "failure_reasons": failure_reasons(
            completed.returncode,
            page_count,
            expected_pages,
            chunk_count,
            body_chunk_count,
            total_chunk_count,
            review_html.exists(),
            excel_check,
        ),
        "stderr_tail": tail(completed.stderr),
    }


def failure_reasons(
    returncode: int,
    page_count: int,
    expected_pages: int,
    chunk_count: int,
    body_chunk_count: int,
    total_chunk_count: int,
    review_exists: bool,
    excel_check: dict[str, Any],
) -> list[str]:
    reasons = []
    if returncode != 0:
        reasons.append(f"main.py exited with {returncode}")
    if page_count != expected_pages:
        reasons.append(f"expected {expected_pages} pages, got {page_count}")
    if chunk_count <= 0:
        reasons.append("summary chunk_count is zero")
    if body_chunk_count <= 0:
        reasons.append("body chunk manifest is empty")
    if total_chunk_count != expected_pages:
        reasons.append(f"expected {expected_pages} total chunks, got {total_chunk_count}")
    if not review_exists:
        reasons.append("review.html missing")
    if not excel_check["exists"]:
        reasons.append("result.xlsx missing")
    if not excel_check["original_table_first"]:
        reasons.append("원본표 is not the first sheet")
    if not excel_check["original_table_non_empty"]:
        reasons.append("원본표 is empty")
    if not excel_check["original_row_coverage_ok"]:
        reasons.append(
            "원본표 row coverage failed "
            f"({excel_check['original_table_row_count']}/{excel_check['source_row_count']})"
        )
    if not excel_check["original_cell_coverage_ok"]:
        reasons.append(
            "원본표 cell coverage failed "
            f"({excel_check['original_table_nonblank_cell_count']}/{excel_check['source_nonblank_cell_count']})"
        )
    if not excel_check["normalized_sheet_exists"]:
        reasons.append("전체명세_정규화 sheet missing")
    if not excel_check["checksum_sheet_exists"]:
        reasons.append("검산 sheet missing")
    return reasons


def inspect_excel(path: Path, original_source: dict[str, Any] | None = None) -> dict[str, Any]:
    original_source = original_source or {"row_count": 0, "nonblank_cell_count": 0}
    result = {
        "exists": path.exists(),
        "sheet_names": [],
        "original_table_first": False,
        "original_table_non_empty": False,
        "original_table_row_count": 0,
        "original_table_column_count": 0,
        "original_table_nonblank_cell_count": 0,
        "source_row_count": int(original_source.get("row_count", 0) or 0),
        "source_nonblank_cell_count": int(original_source.get("nonblank_cell_count", 0) or 0),
        "original_row_coverage_rate": 0.0,
        "original_cell_coverage_rate": 0.0,
        "original_row_coverage_ok": False,
        "original_cell_coverage_ok": False,
        "normalized_sheet_exists": False,
        "checksum_sheet_exists": False,
    }
    if not path.exists():
        return result
    try:
        from openpyxl import load_workbook
    except ImportError:
        return result
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return result
    try:
        result["sheet_names"] = list(workbook.sheetnames)
        result["original_table_first"] = bool(workbook.sheetnames and workbook.sheetnames[0] == "원본표")
        result["normalized_sheet_exists"] = "전체명세_정규화" in workbook.sheetnames
        result["checksum_sheet_exists"] = "검산" in workbook.sheetnames
        if "원본표" in workbook.sheetnames:
            sheet = workbook["원본표"]
            result["original_table_non_empty"] = sheet.max_row > 1 and sheet.max_column > 0
        preservation_sheet_name = "원본표_개발자" if "원본표_개발자" in workbook.sheetnames else "원본표"
        if preservation_sheet_name in workbook.sheetnames:
            sheet = workbook[preservation_sheet_name]
            result["original_table_row_count"] = max(0, sheet.max_row - 1)
            result["original_table_column_count"] = sheet.max_column
            result["original_table_nonblank_cell_count"] = count_original_table_cells(sheet, min_col=5 if preservation_sheet_name == "원본표_개발자" else 1)
    finally:
        workbook.close()
    result["original_row_coverage_rate"] = coverage_rate(
        result["original_table_row_count"], result["source_row_count"]
    )
    result["original_cell_coverage_rate"] = coverage_rate(
        result["original_table_nonblank_cell_count"], result["source_nonblank_cell_count"]
    )
    result["original_row_coverage_ok"] = result["source_row_count"] == 0 or (
        result["original_table_row_count"] >= result["source_row_count"]
    )
    result["original_cell_coverage_ok"] = result["source_nonblank_cell_count"] == 0 or (
        result["original_table_nonblank_cell_count"] >= result["source_nonblank_cell_count"]
    )
    return result


def inspect_original_source(path: Path) -> dict[str, Any]:
    rows = read_jsonl(path)
    nonblank_cell_count = 0
    header_count = 0
    max_cell_count = 0
    for row in rows:
        raw = row.get("raw", {}) if isinstance(row.get("raw"), dict) else {}
        headers = raw.get("header", []) if isinstance(raw.get("header"), list) else []
        cells = raw.get("cells", []) if isinstance(raw.get("cells"), list) else []
        header_count = max(header_count, len(headers))
        max_cell_count = max(max_cell_count, len(cells))
        nonblank_cell_count += sum(1 for cell in cells if str(cell or "").strip())
    return {
        "row_count": len(rows),
        "nonblank_cell_count": nonblank_cell_count,
        "max_header_count": header_count,
        "max_cell_count": max_cell_count,
    }


def count_original_table_cells(sheet: Any, min_col: int = 1) -> int:
    count = 0
    for row in sheet.iter_rows(min_row=2, min_col=min_col, values_only=True):
        count += sum(1 for value in row if str(value or "").strip())
    return count


def coverage_rate(actual: int, expected: int) -> float:
    if expected <= 0:
        return 1.0 if actual >= 0 else 0.0
    return round(actual / expected, 4)


def read_jsonl(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    rows: list[dict[str, Any]] = []
    for line in path.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        try:
            payload = json.loads(line)
        except json.JSONDecodeError:
            continue
        if isinstance(payload, dict):
            rows.append(payload)
    return rows


def manifest_count(path: Path) -> int:
    payload = read_json(path)
    return len(payload) if isinstance(payload, list) else 0


def read_json_object(path: Path) -> dict[str, Any]:
    payload = read_json(path)
    return payload if isinstance(payload, dict) else {}


def read_json(path: Path) -> Any:
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return None


def markdown_report(report: dict[str, Any], root: Path) -> str:
    rows = [
        "# Sample Regression Report",
        "",
        f"- Created: {report['created_at']}",
        f"- Mode: {report['mode']}",
        f"- Samples: {report['sample_count']}",
        f"- Pass: {report['pass_count']}",
        f"- Fail: {report['fail_count']}",
        "",
        "| Sample | Pages | Chunks | Vision | Rows | Original Table | Checksum | Status |",
        "|---|---:|---:|---:|---:|---:|---|---|",
    ]
    for result in report["results"]:
        rows.append(
            "| {sample} | {page_count}/{expected_pages} | {chunk_count} "
            "({body_chunk_count}+{total_chunk_count}) | {vision_ok}/{vision_errors} | "
            "{transaction_count} | {original_table_row_count}/{original_source_row_count} rows, "
            "{original_table_nonblank_cell_count}/{original_source_nonblank_cell_count} cells | "
            "{checksum_status} | {status} |".format(**result)
        )
    failures = [result for result in report["results"] if result["status"] != "PASS"]
    if failures:
        rows.extend(["", "## Failures", ""])
        for result in failures:
            rows.append(f"- `{result['sample']}`: {', '.join(result['failure_reasons'])}")
    rows.extend(["", "## Review Files", ""])
    for result in report["results"]:
        review_path = Path(result["review_html"])
        try:
            review_text = review_path.resolve().relative_to(root.resolve())
        except ValueError:
            review_text = review_path
        rows.append(f"- `{result['sample']}`: `{review_text}`")
    rows.append("")
    return "\n".join(rows)


def tail(value: str | None, max_chars: int = 1200) -> str:
    if value is None:
        return ""
    value = value.strip()
    if len(value) <= max_chars:
        return value
    return value[-max_chars:]


if __name__ == "__main__":
    raise SystemExit(main())
