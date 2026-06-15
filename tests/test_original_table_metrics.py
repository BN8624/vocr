# 원본표 회귀 리포트 지표를 검증하는 테스트
from __future__ import annotations

import json
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import Workbook

from tests.regression_samples import inspect_excel, inspect_original_source


def main() -> int:
    with tempfile.TemporaryDirectory() as temp_dir:
        root = Path(temp_dir)
        source_path = root / "rows_merged.jsonl"
        workbook_path = root / "result.xlsx"
        _write_jsonl(
            source_path,
            [
                {"raw": {"header": ["이용일", "가맹점명"], "cells": ["01.01", "상점A"]}},
                {"raw": {"header": ["이용일", "가맹점명"], "cells": ["01.02", ""]}},
            ],
        )
        _write_workbook(workbook_path)

        source = inspect_original_source(source_path)
        metrics = inspect_excel(workbook_path, source)

        assert source["row_count"] == 2
        assert source["nonblank_cell_count"] == 3
        assert metrics["original_table_first"] is True
        assert metrics["original_table_row_count"] == 2
        assert metrics["original_table_nonblank_cell_count"] == 3
        assert metrics["original_row_coverage_rate"] == 1.0
        assert metrics["original_cell_coverage_rate"] == 1.0
        assert metrics["original_row_coverage_ok"] is True
        assert metrics["original_cell_coverage_ok"] is True

    print("original table metrics test passed")
    return 0


def _write_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "원본표"
    sheet.append(["page", "chunk_id", "local_row_index", "row_type", "이용일", "가맹점명"])
    sheet.append([1, "page_001", 1, "raw", "01.01", "상점A"])
    sheet.append([1, "page_001", 2, "raw", "01.02", ""])
    workbook.create_sheet("전체명세_정규화")
    workbook.create_sheet("검산")
    workbook.save(path)


def _write_jsonl(path: Path, rows: list[dict[str, object]]) -> None:
    with path.open("w", encoding="utf-8") as file:
        for row in rows:
            file.write(json.dumps(row, ensure_ascii=False) + "\n")


if __name__ == "__main__":
    raise SystemExit(main())
