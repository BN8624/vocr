from __future__ import annotations

import json
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import load_workbook

from src.excel_exporter import export_excel
from src.validator import ValidationOutput


def main() -> int:
    with tempfile.TemporaryDirectory() as temp_dir:
        root = Path(temp_dir)
        merged_dir = root / "merged"
        merged_dir.mkdir()
        validated_path = merged_dir / "transactions_validated.jsonl"
        issues_path = merged_dir / "validation_issues.json"
        summary_path = merged_dir / "validation_summary.json"
        _write_jsonl(validated_path, [_row()])
        summary = {
            "checksum": {
                "status": "no_user_total_selected",
                "amount_total": 1000,
                "billing_amount_total": 1000,
                "source_total_candidates": [],
            },
            "column_quality": {
                "issues": [
                    {
                        "code": "card_label_column_contaminated",
                        "field": "card_label",
                        "message": "column value looks suspicious",
                        "value": {"unique_count": 2, "long_text_rate": 0.5},
                        "threshold": {"unique_count": 5, "long_text_rate": 0.4},
                        "header": ["date", "card", "merchant", "amount"],
                    }
                ]
            },
        }
        summary_path.write_text(json.dumps(summary, ensure_ascii=False), encoding="utf-8")
        issues_path.write_text("{}", encoding="utf-8")

        output = export_excel(
            validation_output=ValidationOutput(
                validated_transactions_path=validated_path,
                issues_path=issues_path,
                summary_path=summary_path,
                transaction_count=1,
                row_issue_count=0,
                issue_row_count=0,
                checksum_status="no_user_total_selected",
                checksum_difference=None,
                summary=summary,
            ),
            output_dir=root,
            source_rows_path=validated_path,
        )

        assert output is not None
        assert output.workbook_path.exists()
        workbook = load_workbook(output.workbook_path, read_only=True)
        try:
            assert workbook.sheetnames == ["원본표", "원본표_개발자", "전체명세_정규화", "검산", "원본셀", "추가필드", "확인필요"]
            review_sheet = workbook["확인필요"]
            values = [cell.value for cell in next(review_sheet.iter_rows(min_row=2, max_row=2))]
            assert '{"long_text_rate": 0.5, "unique_count": 2}' in values
        finally:
            workbook.close()

    print("excel exporter test passed")
    return 0


def _row() -> dict[str, object]:
    return {
        "source": {"file": "sample.pdf", "page": 1, "chunk_id": "page_001_chunk_01", "local_row_index": 1},
        "raw": {
            "header": ["date", "card", "merchant", "amount"],
            "cells": ["03.14", "card", "store", "1,000"],
            "image_ref": "chunks/page_001_chunk_01.png",
        },
        "transaction": {
            "date": "03.14",
            "card_label": "card",
            "merchant": "store",
            "amount": 1000,
            "billing_amount": 1000,
            "transaction_type": "",
        },
        "quality": {"needs_review": False, "review_reason": ""},
        "validation": {"needs_review": False, "issues": []},
        "extra_fields": {"raw_metric": {"a": 1}},
    }


def _write_jsonl(path: Path, rows: list[dict[str, object]]) -> None:
    with path.open("w", encoding="utf-8") as file:
        for row in rows:
            file.write(json.dumps(row, ensure_ascii=False) + "\n")


if __name__ == "__main__":
    raise SystemExit(main())
