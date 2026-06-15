from __future__ import annotations

import json
import re
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any

from src.validator import ValidationOutput


@dataclass(frozen=True)
class ExcelExportOutput:
    workbook_path: Path
    sheet_names: list[str]
    transaction_count: int
    review_count: int


def export_excel(
    validation_output: ValidationOutput | None,
    output_dir: Path,
    filename: str = "result.xlsx",
    source_rows_path: Path | None = None,
) -> ExcelExportOutput | None:
    if not validation_output or not validation_output.validated_transactions_path.exists():
        return None

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required for Excel export. Run: pip install -r requirements.txt"
        ) from exc

    rows = _read_jsonl(validation_output.validated_transactions_path)
    raw_rows = _read_jsonl(source_rows_path) if source_rows_path and source_rows_path.exists() else rows
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    summary = validation_output.summary
    ws_original = workbook.create_sheet("원본표")
    ws_original_dev = workbook.create_sheet("원본표_개발자")
    ws_transactions = workbook.create_sheet("전체명세_정규화")
    ws_checksum = workbook.create_sheet("검산")
    ws_raw = workbook.create_sheet("원본셀")
    ws_extra = workbook.create_sheet("추가필드")
    ws_review = workbook.create_sheet("확인필요")

    _write_user_original_table(ws_original, raw_rows, output_dir)
    _write_developer_original_table(ws_original_dev, raw_rows)
    _write_transactions(ws_transactions, rows)
    _write_checksum(ws_checksum, summary)
    _write_raw_cells(ws_raw, raw_rows)
    _write_extra_fields(ws_extra, rows)
    review_count = _write_review_rows(ws_review, rows, summary)

    for index, sheet in enumerate(workbook.worksheets, start=1):
        _style_sheet(sheet, index, get_column_letter, Table, TableStyleInfo, PatternFill, Font, Alignment)

    workbook_path = output_dir / filename
    output_dir.mkdir(parents=True, exist_ok=True)
    workbook.save(workbook_path)

    return ExcelExportOutput(
        workbook_path=workbook_path,
        sheet_names=[sheet.title for sheet in workbook.worksheets],
        transaction_count=len(rows),
        review_count=review_count,
    )


def load_excel_export(output_dir: Path, filename: str = "result.xlsx") -> ExcelExportOutput | None:
    workbook_path = output_dir / filename
    if not workbook_path.exists():
        return None
    return ExcelExportOutput(
        workbook_path=workbook_path,
        sheet_names=[],
        transaction_count=0,
        review_count=0,
    )


def _write_user_original_table(sheet: Any, rows: list[dict[str, Any]], output_dir: Path) -> None:
    headers = _dominant_user_headers(rows)
    sheet.append(headers)
    if not headers:
        return

    rows_for_table = [
        row for row in rows
        if _is_user_row(row) and _same_header(_raw_headers(row), headers)
    ]
    date_column_indexes = [index for index, header in enumerate(headers) if _is_date_header(header)]
    statement_period = _extract_statement_period(output_dir, rows_for_table)
    date_years = _infer_years_for_rows(rows_for_table, date_column_indexes, statement_period)

    for row_index, row in enumerate(rows_for_table):
        raw = _dict(row.get("raw"))
        cells = _list(raw.get("cells"))
        values: list[Any] = []
        for column_index, header in enumerate(headers):
            cell_value = cells[column_index] if column_index < len(cells) else ""
            values.append(_user_cell_value(header, cell_value, date_years.get((row_index, column_index))))
        sheet.append(values)


def _write_developer_original_table(sheet: Any, rows: list[dict[str, Any]]) -> None:
    tracking_headers = ["page", "chunk_id", "local_row_index", "row_type"]
    original_headers, max_extra_count = _collect_original_headers(rows)
    extra_headers = [f"extra_col_{index}" for index in range(1, max_extra_count + 1)]
    sheet.append(tracking_headers + original_headers + extra_headers)

    for row in rows:
        source = _dict(row.get("source"))
        raw = _dict(row.get("raw"))
        row_headers = [str(value) for value in _list(raw.get("header"))]
        cells = ["" if value is None else str(value) for value in _list(raw.get("cells"))]
        value_by_header: dict[str, str] = {}
        extras: list[str] = []
        for index, cell in enumerate(cells):
            if index < len(row_headers):
                header = row_headers[index]
                if header and header not in value_by_header:
                    value_by_header[header] = cell
                    continue
            extras.append(cell)

        row_type = _original_row_type(row)
        sheet.append(
            [
                source.get("page", ""),
                source.get("chunk_id", ""),
                source.get("local_row_index", ""),
                row_type,
            ]
            + [value_by_header.get(header, "") for header in original_headers]
            + [extras[index] if index < len(extras) else "" for index in range(max_extra_count)]
        )


def _dominant_user_headers(rows: list[dict[str, Any]]) -> list[str]:
    counts: dict[tuple[str, ...], int] = {}
    first_seen: dict[tuple[str, ...], int] = {}
    headers_by_key: dict[tuple[str, ...], list[str]] = {}
    for index, row in enumerate(rows):
        if not _is_user_row(row):
            continue
        headers = _raw_headers(row)
        if not headers:
            continue
        key = tuple(_normalize_header(header) for header in headers)
        counts[key] = counts.get(key, 0) + 1
        first_seen.setdefault(key, index)
        headers_by_key.setdefault(key, headers)
    if not counts:
        return []
    selected = max(counts, key=lambda key: (counts[key], -first_seen[key]))
    return headers_by_key[selected]


def _is_user_row(row: dict[str, Any]) -> bool:
    row_type = str(row.get("row_type") or "").strip().lower()
    if row_type in {"total", "section", "note"}:
        return False
    raw = _dict(row.get("raw"))
    cells = _list(raw.get("cells"))
    return any(str(cell or "").strip() for cell in cells)


def _raw_headers(row: dict[str, Any]) -> list[str]:
    raw = _dict(row.get("raw"))
    return [str(value) for value in _list(raw.get("header"))]


def _same_header(left: list[str], right: list[str]) -> bool:
    return [_normalize_header(value) for value in left] == [_normalize_header(value) for value in right]


def _normalize_header(value: str) -> str:
    return re.sub(r"\s+", "", str(value or "")).strip()


def _collect_original_headers(rows: list[dict[str, Any]]) -> tuple[list[str], int]:
    headers: list[str] = []
    seen: set[str] = set()
    max_extra_count = 0
    for row in rows:
        raw = _dict(row.get("raw"))
        row_headers = [str(value) for value in _list(raw.get("header"))]
        cells = _list(raw.get("cells"))
        for header in row_headers:
            if header and header not in seen:
                seen.add(header)
                headers.append(header)
        max_extra_count = max(max_extra_count, max(0, len(cells) - len(row_headers)))
    return headers, max_extra_count


def _original_row_type(row: dict[str, Any]) -> str:
    if row.get("row_type"):
        return str(row.get("row_type"))
    merge = _dict(row.get("merge"))
    if merge.get("is_duplicate"):
        return "duplicate"
    quality = _dict(row.get("quality"))
    validation = _dict(row.get("validation"))
    if quality.get("needs_review") or validation.get("needs_review"):
        return "needs_review"
    transaction = _dict(row.get("transaction"))
    if transaction:
        return "transaction"
    return "raw"


def _write_transactions(sheet: Any, rows: list[dict[str, Any]]) -> None:
    headers = [
        "source_file",
        "page",
        "chunk_id",
        "local_row_index",
        "date",
        "card_label",
        "merchant",
        "transaction_type",
        "amount",
        "billing_amount",
        "needs_review",
        "review_reason",
    ]
    sheet.append(headers)
    for row in rows:
        source = _dict(row.get("source"))
        transaction = _dict(row.get("transaction"))
        quality = _dict(row.get("quality"))
        validation = _dict(row.get("validation"))
        needs_review = bool(quality.get("needs_review") or validation.get("needs_review"))
        sheet.append(
            [
                source.get("file", ""),
                source.get("page", ""),
                source.get("chunk_id", ""),
                source.get("local_row_index", ""),
                transaction.get("date", ""),
                transaction.get("card_label", ""),
                transaction.get("merchant", ""),
                transaction.get("transaction_type", ""),
                _number_or_blank(transaction.get("amount")),
                _number_or_blank(transaction.get("billing_amount")),
                "Y" if needs_review else "",
                quality.get("review_reason", ""),
            ]
        )


def _write_checksum(sheet: Any, summary: dict[str, Any]) -> None:
    checksum = _dict(summary.get("checksum"))
    sheet.append(["item", "value"])
    sheet.append(["status", checksum.get("status", "")])
    sheet.append(["message", checksum.get("message", "")])
    sheet.append(["amount_total", checksum.get("amount_total", "")])
    sheet.append(["billing_amount_total", checksum.get("billing_amount_total", "")])
    selected = checksum.get("selected_total", {}) if isinstance(checksum.get("selected_total"), dict) else {}
    sheet.append(["selected_total_id", checksum.get("selected_total_id", "")])
    sheet.append(["selected_total_label", selected.get("label", "")])
    sheet.append(["selected_total_amount", selected.get("amount", "")])
    sheet.append(["difference", checksum.get("difference", "")])
    sheet.append(["processed_chunk_count", checksum.get("processed_chunk_count", "")])
    sheet.append(["expected_chunk_count", checksum.get("expected_chunk_count", "")])
    sheet.append([])
    sheet.append(["source_total_label", "value_text", "amount", "page", "chunk_id", "needs_review", "review_reason"])
    for candidate in checksum.get("source_total_candidates", []) or []:
        if not isinstance(candidate, dict):
            continue
        sheet.append(
            [
                candidate.get("label", ""),
                candidate.get("value_text", ""),
                candidate.get("amount", ""),
                candidate.get("page", ""),
                candidate.get("chunk_id", ""),
                "Y" if candidate.get("needs_review") else "",
                candidate.get("review_reason", ""),
            ]
        )
    if not checksum.get("source_total_candidates"):
        sheet.append(["원본 합계 후보 없음", "", "", "", "", "", ""])
    auto_matches = [item for item in checksum.get("auto_match_candidates", []) or [] if isinstance(item, dict)]
    if auto_matches:
        sheet.append([])
        sheet.append(["auto_match_field", "source_total_label", "amount", "chunk_id"])
        for item in auto_matches:
            candidate = item.get("candidate", {}) if isinstance(item.get("candidate"), dict) else {}
            sheet.append(
                [
                    item.get("field", ""),
                    candidate.get("label", ""),
                    candidate.get("amount", ""),
                    candidate.get("chunk_id", ""),
                ]
            )


def _write_raw_cells(sheet: Any, rows: list[dict[str, Any]]) -> None:
    sheet.append(
        [
            "source_file",
            "page",
            "chunk_id",
            "local_row_index",
            "cell_index",
            "header",
            "cell_value",
            "image_ref",
        ]
    )
    for row in rows:
        source = _dict(row.get("source"))
        raw = _dict(row.get("raw"))
        headers = [str(value) for value in raw.get("header", [])]
        cells = [str(value) for value in raw.get("cells", [])]
        for index, cell in enumerate(cells, start=1):
            header = headers[index - 1] if index <= len(headers) else f"col_{index}"
            sheet.append(
                [
                    source.get("file", ""),
                    source.get("page", ""),
                    source.get("chunk_id", ""),
                    source.get("local_row_index", ""),
                    index,
                    header,
                    cell,
                    raw.get("image_ref", ""),
                ]
            )


def _write_extra_fields(sheet: Any, rows: list[dict[str, Any]]) -> None:
    sheet.append(["source_file", "page", "chunk_id", "local_row_index", "field", "value"])
    for row in rows:
        source = _dict(row.get("source"))
        extra_fields = _dict(row.get("extra_fields"))
        for key, value in extra_fields.items():
            values = value if isinstance(value, list) else [value]
            for item in values:
                sheet.append(
                    [
                        source.get("file", ""),
                        source.get("page", ""),
                        source.get("chunk_id", ""),
                        source.get("local_row_index", ""),
                        key,
                        _excel_value(item),
                    ]
                )


def _write_review_rows(sheet: Any, rows: list[dict[str, Any]], summary: dict[str, Any]) -> int:
    column_quality = _dict(summary.get("column_quality"))
    column_issues = [
        issue for issue in column_quality.get("issues", []) if isinstance(issue, dict)
    ]
    if column_issues:
        sheet.append(["type", "code", "field", "message", "value", "threshold", "header"])
        for issue in column_issues:
            sheet.append(
                [
                    "column_quality",
                    issue.get("code", ""),
                    issue.get("field", ""),
                    issue.get("message", ""),
                    _excel_value(issue.get("value", "")),
                    _excel_value(issue.get("threshold", "")),
                    " | ".join(str(value) for value in issue.get("header", [])),
                ]
            )
        sheet.append([])

    sheet.append(
        [
            "source_file",
            "page",
            "chunk_id",
            "local_row_index",
            "date",
            "merchant",
            "amount",
            "quality_reason",
            "validation_issues",
            "raw_cells",
            "image_ref",
        ]
    )
    review_count = len(column_issues)
    for row in rows:
        quality = _dict(row.get("quality"))
        validation = _dict(row.get("validation"))
        issues = [issue for issue in validation.get("issues", []) if isinstance(issue, dict)]
        if not quality.get("needs_review") and not validation.get("needs_review") and not issues:
            continue
        review_count += 1
        source = _dict(row.get("source"))
        raw = _dict(row.get("raw"))
        transaction = _dict(row.get("transaction"))
        sheet.append(
            [
                source.get("file", ""),
                source.get("page", ""),
                source.get("chunk_id", ""),
                source.get("local_row_index", ""),
                transaction.get("date", ""),
                transaction.get("merchant", ""),
                _number_or_blank(transaction.get("amount")),
                quality.get("review_reason", ""),
                "; ".join(str(issue.get("message", "")) for issue in issues),
                " | ".join(str(value) for value in raw.get("cells", [])),
                raw.get("image_ref", ""),
            ]
        )
    return review_count


def _style_sheet(
    sheet: Any,
    sheet_index: int,
    get_column_letter: Any,
    Table: Any,
    TableStyleInfo: Any,
    PatternFill: Any,
    Font: Any,
    Alignment: Any,
) -> None:
    sheet.freeze_panes = "A2"
    max_row = sheet.max_row
    max_col = sheet.max_column
    header_fill = PatternFill("solid", fgColor="176B5D")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        if cell.value is None:
            cell.value = f"column_{cell.column}"
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for column_index in range(1, max_col + 1):
        letter = get_column_letter(column_index)
        values = [sheet.cell(row=row_index, column=column_index).value for row_index in range(1, min(max_row, 80) + 1)]
        width = min(42, max(10, max(len(str(value)) if value is not None else 0 for value in values) + 2))
        sheet.column_dimensions[letter].width = width

    for row_index in range(2, max_row + 1):
        if row_index % 2 == 0:
            for cell in sheet[row_index]:
                cell.fill = PatternFill("solid", fgColor="F7FAF8")

    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, int):
                cell.number_format = '#,##0'
            elif isinstance(cell.value, date):
                cell.number_format = 'yyyy-mm-dd'

    if max_row >= 2 and max_col >= 1:
        table_ref = f"A1:{get_column_letter(max_col)}{max_row}"
        table_name = _safe_table_name(sheet_index)
        table = Table(displayName=table_name, ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)


def _safe_table_name(sheet_index: int) -> str:
    return f"Table{sheet_index}"


def _user_cell_value(header: str, value: Any, inferred_year: int | None = None) -> Any:
    text = "" if value is None else str(value).strip()
    if not text:
        return ""
    if _is_date_header(header):
        parsed = _parse_date_value(text, inferred_year)
        return parsed if parsed else text
    if _is_amount_header(header):
        parsed_number = _parse_number_value(text)
        return parsed_number if parsed_number is not None else text
    return text


def _is_date_header(header: str) -> bool:
    normalized = _normalize_header(header)
    return normalized in {"이용일", "거래일", "사용일", "승인일", "매출일", "date"}


def _is_amount_header(header: str) -> bool:
    normalized = _normalize_header(header)
    amount_keywords = ("금액", "원금", "포인트", "point", "amount", "합계")
    if "현지" in normalized and "금액" in normalized:
        return True
    return any(keyword in normalized.lower() for keyword in amount_keywords)


def _parse_number_value(value: str) -> int | float | None:
    text = value.strip().replace(",", "")
    if not text:
        return None
    if text.startswith("(") and text.endswith(")"):
        text = f"-{text[1:-1]}"
    if text.endswith("-"):
        text = f"-{text[:-1]}"
    text = re.sub(r"[^0-9.\-]", "", text)
    if not text or text in {"-", ".", "-."}:
        return None
    try:
        number = float(text) if "." in text else int(text)
    except ValueError:
        return None
    if isinstance(number, float) and number.is_integer():
        return int(number)
    return number


def _parse_date_value(value: str, inferred_year: int | None = None) -> date | None:
    text = value.strip()
    match = re.fullmatch(r"(\d{4})[.\-/년\s]*(\d{1,2})[.\-/월\s]*(\d{1,2})일?", text)
    if match:
        return _safe_date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
    match = re.fullmatch(r"(\d{2})[.\-/](\d{1,2})[.\-/](\d{1,2})", text)
    if match:
        return _safe_date(2000 + int(match.group(1)), int(match.group(2)), int(match.group(3)))
    match = re.fullmatch(r"(\d{1,2})[.\-/](\d{1,2})", text)
    if match and inferred_year:
        return _safe_date(inferred_year, int(match.group(1)), int(match.group(2)))
    match = re.fullmatch(r"(\d{2})(\d{2})", text)
    if match and inferred_year:
        return _safe_date(inferred_year, int(match.group(1)), int(match.group(2)))
    return None


def _safe_date(year: int, month: int, day: int) -> date | None:
    try:
        return date(year, month, day)
    except ValueError:
        return None


def _infer_years_for_rows(
    rows: list[dict[str, Any]],
    date_column_indexes: list[int],
    statement_period: tuple[date, date] | None = None,
) -> dict[tuple[int, int], int]:
    inferred: dict[tuple[int, int], int] = {}
    if not date_column_indexes:
        return inferred
    explicit_year = statement_period[0].year if statement_period else _first_explicit_year(rows, date_column_indexes)
    parsed_months = _date_tokens(rows, date_column_indexes)
    saw_wrap = any(
        left[2] >= 10 and right[2] <= 2
        for left, right in zip(parsed_months, parsed_months[1:])
    )
    base_year = explicit_year if explicit_year is not None else date.today().year
    current_year = base_year - 1 if explicit_year is None and saw_wrap else base_year
    if statement_period and parsed_months:
        start, end = statement_period
        first_month, first_day = parsed_months[0][2], parsed_months[0][3]
        end_year_candidate = _safe_date(end.year, first_month, first_day)
        if end_year_candidate and start <= end_year_candidate <= end:
            current_year = end.year
    previous_month: int | None = None
    for row_index, row in enumerate(rows):
        cells = _list(_dict(row.get("raw")).get("cells"))
        for column_index in date_column_indexes:
            value = cells[column_index] if column_index < len(cells) else ""
            month_day = _parse_month_day(str(value or ""))
            if not month_day:
                continue
            month, _day = month_day
            if previous_month is not None and previous_month >= 10 and month <= 2:
                current_year += 1
            inferred[(row_index, column_index)] = current_year
            previous_month = month
    return inferred


def _date_tokens(rows: list[dict[str, Any]], date_column_indexes: list[int]) -> list[tuple[int, int, int, int]]:
    tokens: list[tuple[int, int, int, int]] = []
    for row_index, row in enumerate(rows):
        cells = _list(_dict(row.get("raw")).get("cells"))
        for column_index in date_column_indexes:
            value = cells[column_index] if column_index < len(cells) else ""
            month_day = _parse_month_day(str(value or ""))
            if month_day:
                tokens.append((row_index, column_index, month_day[0], month_day[1]))
    return tokens


def _first_explicit_year(rows: list[dict[str, Any]], date_column_indexes: list[int]) -> int | None:
    for row in rows:
        cells = _list(_dict(row.get("raw")).get("cells"))
        for column_index in date_column_indexes:
            value = str(cells[column_index] if column_index < len(cells) else "")
            match = re.search(r"(\d{4})[.\-/년\s]*\d{1,2}[.\-/월\s]*\d{1,2}", value)
            if match:
                return int(match.group(1))
            match = re.search(r"(\d{2})[.\-/]\d{1,2}[.\-/]\d{1,2}", value)
            if match:
                return 2000 + int(match.group(1))
    return None


def _parse_month_day(value: str) -> tuple[int, int] | None:
    text = value.strip()
    match = re.fullmatch(r"\d{4}[.\-/년\s]*(\d{1,2})[.\-/월\s]*(\d{1,2})일?", text)
    if match:
        return int(match.group(1)), int(match.group(2))
    match = re.fullmatch(r"\d{2}[.\-/](\d{1,2})[.\-/](\d{1,2})", text)
    if match:
        return int(match.group(1)), int(match.group(2))
    match = re.fullmatch(r"(\d{1,2})[.\-/](\d{1,2})", text)
    if match:
        return int(match.group(1)), int(match.group(2))
    match = re.fullmatch(r"(\d{2})(\d{2})", text)
    if match:
        return int(match.group(1)), int(match.group(2))
    return None


def _extract_statement_period(output_dir: Path, rows: list[dict[str, Any]]) -> tuple[date, date] | None:
    image_path = _first_page_image_path(output_dir, rows)
    if not image_path or not image_path.exists():
        return None
    try:
        from PIL import Image
        import easyocr
    except Exception:
        return None
    try:
        image = Image.open(image_path)
        width, height = image.size
        crop = image.crop((int(width * 0.4), int(height * 0.08), int(width * 0.98), int(height * 0.24)))
        temp_path = output_dir / "merged" / "statement_period_crop.png"
        temp_path.parent.mkdir(parents=True, exist_ok=True)
        crop.save(temp_path)
        reader = easyocr.Reader(["ko", "en"], gpu=False, verbose=False)
        text = " ".join(str(item) for item in reader.readtext(str(temp_path), detail=0))
    except Exception:
        return None
    return _parse_statement_period_text(text)


def _first_page_image_path(output_dir: Path, rows: list[dict[str, Any]]) -> Path | None:
    for row in rows:
        raw = _dict(row.get("raw"))
        image_ref = str(raw.get("image_ref") or "").strip()
        if image_ref:
            return output_dir / image_ref
    return None


def _parse_statement_period_text(text: str) -> tuple[date, date] | None:
    normalized = re.sub(r"(?<=\d),(?=\d{1,2}\b)", ".", text)
    normalized = re.sub(r"\s+", " ", normalized)
    matches = re.findall(r"(20\d{2})\D{0,3}(\d{1,2})\D{0,3}(\d{1,2})", normalized)
    dates = [_safe_date(int(year), int(month), int(day)) for year, month, day in matches]
    dates = [value for value in dates if value is not None]
    if len(dates) < 2:
        return None
    start, end = dates[0], dates[1]
    if start <= end:
        return start, end
    return None


def _number_or_blank(value: Any) -> int | str:
    return value if isinstance(value, int) else ""


def _excel_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    return json.dumps(value, ensure_ascii=False, sort_keys=True)


def _dict(value: Any) -> dict[str, Any]:
    return value if isinstance(value, dict) else {}


def _list(value: Any) -> list[Any]:
    return value if isinstance(value, list) else []


def _read_jsonl(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for line in path.read_text(encoding="utf-8").splitlines():
        if line.strip():
            rows.append(json.loads(line))
    return rows
